#!/usr/bin/env python3
"""Generate an Excel file of the PetClinic owners with the most pets."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

rows = [
    (11, "George", "Darling", "14 Kensington Gardens", "London", "0442079372121", 2, "Liza (cat), Nana (dog)"),
    (8, "Alice", "Liddell", "Christ Church", "Oxford", "0441865276150", 2, "Cheshire (cat), Dinah (cat)"),
    (15, "Beatrix", "Potter", "Hill Top Farm", "Near Sawrey", "0441539436269", 2, "Mittens (cat), Pickles (dog)"),
    (6, "Roger", "Radcliff", "27 Outer Circle", "London", "0442074860707", 2, "Perdita (dog), Pongo (dog)"),
    (24, "Lady", "Tremaine", "Chateau Tremaine", "Ile-de-France", "0146203030", 2, "Jaq (hamster), Lucifer (cat)"),
    (18, "Wallace", "Wensleydale", "62 West Wallaby Street", "Wigan", "0441942244466", 2, "Gromit (dog), Hutch (hamster)"),
]

headers = ["Owner ID", "First Name", "Last Name", "Address", "City", "Telephone", "Pet Count", "Pets"]

wb = Workbook()
ws = wb.active
ws.title = "Top Owners"

green = PatternFill(start_color="6DB33F", end_color="6DB33F", fill_type="solid")
white_bold = Font(color="FFFFFF", bold=True)
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.append(headers)
for c in ws[1]:
    c.fill = green
    c.font = white_bold
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

for r in rows:
    ws.append(r)

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for c in row:
        c.border = border
    row[6].alignment = Alignment(horizontal="center")  # Pet Count

widths = [9, 12, 14, 24, 14, 16, 10, 34]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + i)].width = w

ws.freeze_panes = "A2"

out = "top_owners_by_pets.xlsx"
wb.save(out)
print("wrote", out, "with", len(rows), "owners")
